import click
from flask.cli import with_appcontext
from flask.globals import current_app
from openpyxl import load_workbook
from db_models import db, Manuscript, Page
import asyncio
import sys
import aiohttp


@click.command()
@with_appcontext
def create():
    db.create_all()


@click.command()
@click.argument('filename')
@click.option('--sheet', default=1, help='Sheet to use in a multisheet file')
@with_appcontext
def ingest(filename, sheet):
    def get_sheet():
        wb = load_workbook(filename)
        return wb.get_sheet_by_name(wb.get_sheet_names()[sheet-1])

    def clean_decimal(value):
        if type(value) == float:
            return int(value)
        return value

    created_manuscripts = updated_manuscripts = 0
    created_pages = updated_pages = 0

    async def process_row(row):
        nonlocal created_manuscripts, updated_manuscripts
        try:
            heb_name = row[0].value
            iiif_manifest_url = row[5].value
            if not iiif_manifest_url:
                return

            manuscript = Manuscript.query.filter_by(heb_name=heb_name).first()
            if not manuscript:
                created_manuscripts += 1
                manuscript = Manuscript()
                manuscript.heb_name = heb_name
                db.session.add(manuscript)
            else:
                updated_manuscripts += 1
            manuscript.eng_name = None  # No english name for now
            manuscript.estimated_date = clean_decimal(row[1].value)
            manuscript.shelf_no = clean_decimal(row[2].value)
            manuscript.catalog_no = clean_decimal(row[3].value)
            manuscript.external_url = row[4].value
            manuscript.iiif_manifest_url = iiif_manifest_url
            manuscript.notes = row[8].value

            await process_manifest(manuscript)
        except Exception as e:
            print("Can't process manifest: ", e, file=sys.stderr)

    async def process_manifest(manuscript):
        nonlocal created_pages, updated_pages
        async with aiohttp.ClientSession() as session:
            async with session.get(manuscript.iiif_manifest_url) as response:
                manifest = await response.json()
        for i, canvas in enumerate(manifest['sequences'][0]['canvases']):
            url = canvas['@id']
            page = Page.query.filter_by(iiif_url=url).first()
            if page:
                if page.manuscript_id != manuscript.id:
                    raise ValueError(
                        f'Page for {url} found in manuscript {manuscript.heb_name}, but aready belongs to {page.manuscript.heb_name}')
                updated_pages += 1
            else:
                page = Page()
                page.manuscript_id = manuscript.id
                page.iiif_url = url
                db.session.add(page)
                created_pages += 1
            page.page_number = i + 1
            page.page_name = canvas['label']
            page.page_width = canvas['width']
            page.page_height = canvas['height']

    async def process_sheet(sheet):
        row_tasks = [process_row(row) for row in sheet.iter_rows(min_row=2)]
        await asyncio.wait(row_tasks)

    sheet = get_sheet()
    asyncio.run(process_sheet(sheet))
    db.session.commit()
    print(f'Created {created_manuscripts} manuscripts and {created_pages} pages. Updated {updated_manuscripts} manuscripts and {updated_pages} pages')


def init_app(app):
    app.cli.add_command(ingest)
    app.cli.add_command(create)
